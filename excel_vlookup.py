import os
import tkinter as tk
from tkinter import filedialog, ttk, messagebox, Listbox, Scrollbar
from ttkthemes import ThemedTk
import openpyxl
import pandas as pd
import logging
import datetime

class TextHandler(logging.Handler):
    def __init__(self, text_widget):
        logging.Handler.__init__(self)
        self.text_widget = text_widget

    def emit(self, record):
        msg = self.format(record)
        self.text_widget.config(state=tk.NORMAL)
        self.text_widget.insert(tk.END, msg + "\n")
        self.text_widget.see(tk.END)
        self.text_widget.config(state=tk.DISABLED)

class VlookupView:
    def __init__(self, controller):
        self.controller = controller
        self.root = ThemedTk(theme="arc")
        self.root.title("VLOOKUP 工具")
        self.main_saved_selection = []
        self.lookup_index_saved_selection = []
        self.lookup_match_saved_selection = []
        self._build_gui()

    def _build_gui(self):
        self.root.title("高级VLOOKUP工具")

        # 使用 ttkthemes 来应用一个现代化的主题
        style = ttk.Style(self.root)
        style.theme_use('clam')  # 使用 'classic' 主题

        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 主框架的左侧部分
        main_frame_left = ttk.LabelFrame(main_frame, text="主表", padding="10")
        main_frame_left.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")

        ttk.Label(main_frame_left, text="选择文件：").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.main_file_entry = ttk.Entry(main_frame_left, width=20)
        self.main_file_entry.grid(row=0, column=1, pady=5)
        ttk.Button(main_frame_left, text="浏览",
                   command=lambda: self.controller.load_file(self.main_file_entry,
                                                             self.main_sheet_combobox,
                                                             self.main_column_listbox)).grid(row=0, column=2, padx=5)
        ttk.Label(main_frame_left, text="sheet页：").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.main_sheet_combobox = ttk.Combobox(main_frame_left, width=18)
        self.main_sheet_combobox.grid(row=1, column=1, pady=5)

        ttk.Label(main_frame_left, text="索引列：").grid(row=2, column=0, sticky=tk.N, pady=5)
        self.main_column_listbox = Listbox(main_frame_left, selectmode=tk.MULTIPLE, height=5, width=21)
        self.main_column_listbox.grid(row=2, column=1, sticky=tk.W, pady=5)
        self.main_scrollbar = Scrollbar(main_frame_left, command=self.main_column_listbox.yview)
        self.main_column_listbox.config(yscrollcommand=self.main_scrollbar.set)
        self.main_scrollbar.grid(row=2, column=2, sticky=tk.W + tk.N + tk.S, pady=5)
        self.main_column_lock_button = ttk.Button(main_frame_left, text="确认选择",
                                                  command=lambda: self.controller.toggle_lock_columns(
                                                      self.main_column_listbox,
                                                      self.main_column_lock_button,
                                                      self.main_saved_selection))
        self.main_column_lock_button.grid(row=3, column=1, pady=5)
        self.log_text = tk.Text(main_frame_left, height=10, width=50)
        self.log_text.grid(row=4, column=0, columnspan=3, pady=10)
        self.log_text.insert(tk.END, "日志记录：\n")
        self.log_text.config(state=tk.DISABLED)

        # 主框架的右侧部分
        main_frame_right = ttk.LabelFrame(main_frame, text="查找表", padding="10")
        main_frame_right.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")

        ttk.Label(main_frame_right, text="选择文件：").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.lookup_file_entry = ttk.Entry(main_frame_right, width=20)
        self.lookup_file_entry.grid(row=0, column=1, pady=5)
        ttk.Button(main_frame_right, text="浏览",
                   command=lambda: self.controller.load_file(self.lookup_file_entry,
                                                             self.lookup_sheet_combobox,
                                                             self.lookup_index_listbox,
                                                             self.lookup_match_listbox)).grid(row=0, column=2, padx=5)
        ttk.Label(main_frame_right, text="Sheet页：").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.lookup_sheet_combobox = ttk.Combobox(main_frame_right, width=18)
        self.lookup_sheet_combobox.grid(row=1, column=1, pady=5)

        ttk.Label(main_frame_right, text="索引列：").grid(row=2, column=0, sticky=tk.N, pady=5)
        self.lookup_index_listbox = Listbox(main_frame_right, selectmode=tk.MULTIPLE, height=5, width=22)
        self.lookup_index_listbox.grid(row=2, column=1, sticky=tk.W, pady=5)
        self.lookup_scrollbar_index = Scrollbar(main_frame_right, command=self.lookup_index_listbox.yview)
        self.lookup_index_listbox.config(yscrollcommand=self.lookup_scrollbar_index.set)
        self.lookup_scrollbar_index.grid(row=2, column=2, sticky=tk.W + tk.N + tk.S, pady=5)

        ttk.Label(main_frame_right, text="匹配列：").grid(row=4, column=0, sticky=tk.N, padx=10, pady=5)
        self.lookup_match_listbox = Listbox(main_frame_right, selectmode=tk.MULTIPLE, height=5, width=22)
        self.lookup_match_listbox.grid(row=4, column=1, sticky=tk.W, pady=5)
        self.lookup_scrollbar_match = Scrollbar(main_frame_right, command=self.lookup_match_listbox.yview)
        self.lookup_match_listbox.config(yscrollcommand=self.lookup_scrollbar_match.set)
        self.lookup_scrollbar_match.grid(row=4, column=2, sticky=tk.W + tk.N + tk.S, pady=5)

        self.lookup_index_lock_button = ttk.Button(main_frame_right, text="确认选择",
                                                   command=lambda: self.controller.toggle_lock_columns(
                                                       self.lookup_index_listbox,
                                                       self.lookup_index_lock_button,
                                                       self.lookup_index_saved_selection))
        self.lookup_index_lock_button.grid(row=3, column=1, pady=5)

        self.lookup_match_lock_button = ttk.Button(main_frame_right, text="确认选择",
                                                   command=lambda: self.controller.toggle_lock_columns(
                                                       self.lookup_match_listbox,
                                                       self.lookup_match_lock_button,
                                                       self.lookup_match_saved_selection))
        self.lookup_match_lock_button.grid(row=5, column=1, pady=5)

        ttk.Button(main_frame, text="执行VLOOKUP",
                   command=lambda: self.controller.execute_vlookup(self.main_file_entry.get(),
                                                                   self.lookup_file_entry.get(),
                                                                   self.main_sheet_combobox.get(),
                                                                   self.lookup_sheet_combobox.get(),
                                                                   self.main_saved_selection,
                                                                   self.lookup_index_saved_selection,
                                                                   self.lookup_match_saved_selection)).grid(row=1,
                                                                                                            column=0,
                                                                                                            columnspan=2,
                                                                                                            pady=20)

    def update_log(self, message):
        self.log_text.config(state=tk.NORMAL)
        formatted_message = f"{datetime.datetime.now()} - {message}"
        self.log_text.insert(tk.END, formatted_message + "\n")
        self.log_text.see(tk.END)  # 滚动到日志的最后一行
        self.log_text.config(state=tk.DISABLED)

    def run(self):
        self.root.mainloop()


class VlookupController:
    def __init__(self):
        self.model = VlookupModel()
        self.view = VlookupView(self)
        # 初始化日志记录
        self.logger = logging.getLogger('VLOOKUP_TOOL')
        self.logger.setLevel(logging.INFO)

        # 创建处理器并添加到logger
        text_handler = TextHandler(self.view.log_text)
        formatter = logging.Formatter('%(asctime)s - %(message)s')
        text_handler.setFormatter(formatter)
        self.logger.addHandler(text_handler)

    def load_file(self, entry_widget, sheet_combobox, index_listbox, match_listbox=None):
        file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xls;*.xlsx")])
        self.logger.info(f"加载文件: {file_path}")
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, file_path)
        sheets = self.model.get_sheets(file_path)
        sheet_combobox["values"] = sheets
        sheet_combobox.current(0)
        columns = self.model.get_columns(file_path, sheet_combobox.get())
        index_listbox.delete(0, tk.END)
        for column in columns:
            index_listbox.insert(tk.END, column)
        if match_listbox:
            match_listbox.delete(0, tk.END)
            for column in columns:
                match_listbox.insert(tk.END, column)



    def toggle_lock_columns(self, listbox, button, saved_selection):
        if listbox["state"] == tk.NORMAL:
            listbox.config(state=tk.DISABLED)
            button.config(text="确认选择")

            saved_selection[:] = [listbox.get(i) for i in listbox.curselection()]
            self.logger.info(f"选择的列名: {saved_selection[:]}")
        else:
            listbox.config(state=tk.NORMAL)
            button.config(text="取消选择")
            self.logger.info('取消选择')
    def execute_vlookup(self, main_file, lookup_file, main_sheet, lookup_sheet, main_columns, lookup_index_columns,
                        lookup_match_columns):
        self.logger.info("执行VLOOKUP...")

        try:
            result_file_path = self.model.perform_vlookup(main_file, lookup_file, main_sheet, lookup_sheet,
                                                          main_columns, lookup_index_columns, lookup_match_columns)
            self.logger.info( f"VLOOKUP 完成。结果已保存至 {result_file_path}")
            messagebox.showinfo("成功", f"VLOOKUP 完成。结果已保存至 {result_file_path}")
        except Exception as e:
            self.logger.info(f"发生错误: {str(e)}")
            messagebox.showerror("错误", f"发生错误: {str(e)}")

    def run(self):
        self.view.run()


class VlookupModel:
    def get_sheets(self, file_path):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames

    def get_columns(self, file_path, sheet_name):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb[sheet_name]
        first_row = sheet[1]
        return [cell.value for cell in first_row]

    def perform_vlookup(self, main_file, lookup_file, main_sheet, lookup_sheet, main_columns, lookup_index_columns,
                        lookup_match_columns):
        main_df = pd.read_excel(main_file, sheet_name=main_sheet)
        lookup_df = pd.read_excel(lookup_file, sheet_name=lookup_sheet)

        result_df = main_df.merge(lookup_df[lookup_index_columns + lookup_match_columns],
                                  left_on=main_columns,
                                  right_on=lookup_index_columns,
                                  how='left')

        main_file_name = os.path.basename(main_file)
        base_name, ext = os.path.splitext(main_file_name)
        counter = 1
        while True:
            if counter == 1:
                result_file_name = f"{base_name}_Vlookup{ext}"
            else:
                result_file_name = f"{base_name}_Vlookup_{counter}{ext}"

            result_file_path = os.path.join(os.path.dirname(main_file), result_file_name)
            if not os.path.exists(result_file_path):
                break
            counter += 1

        result_df.to_excel(result_file_path, index=False)
        return result_file_path


if __name__ == "__main__":
    app = VlookupController()
    app.run()
